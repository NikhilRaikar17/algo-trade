"""
Daily backtest P&L email report — sent at 4:00 PM IST.

Mon–Thu  →  Today's trades only.
           Excel: Summary sheet (today) + one sheet per strategy (today's trades only).

Friday   →  Full week (up to 5 trading days).
           Excel: Weekly Summary sheet + one sheet per trading day (all strategies).

Instruments: Top active stocks from DB (up to 20), not NIFTY index.
"""

import io
import os
import smtplib
import traceback
from collections import defaultdict
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import now_ist, _is_trading_day
from state import _is_already_sent, _mark_sent
from strategy_registry import get_strategy_short_names

# ── Colour palette ─────────────────────────────────────────────────────────────
_GREEN  = "FF1E8449"
_RED    = "FFC0392B"
_DARK   = "FF2C3E50"
_WHITE  = "FFFFFFFF"
_HEADER_FILL  = PatternFill("solid", fgColor="FF1A5276")
_SUB_FILL     = PatternFill("solid", fgColor="FF2E86C1")
_ALT_FILL     = PatternFill("solid", fgColor="FFEAF4FB")
_WIN_FILL     = PatternFill("solid", fgColor="FFD5F5E3")
_LOSS_FILL    = PatternFill("solid", fgColor="FFFDEDEC")
_SUMMARY_FILL = PatternFill("solid", fgColor="FF1B2631")
_WHITE_FILL   = PatternFill("solid", fgColor="FFFFFFFF")
_PURPLE_FILL  = PatternFill("solid", fgColor="FFE8DAEF")

_THIN   = Side(style="thin", color="FFB2BABB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALL_STRATEGIES = get_strategy_short_names()


# ── Data fetching ──────────────────────────────────────────────────────────────

def _fetch_all_stocks_for_report():
    """Fetch candles for all active top stocks and run all backtests. Returns merged trades."""
    from db import get_active_top_stocks                               # noqa: PLC0415
    from pages.backtest_pnl_tab import _fetch_all_stocks_trades       # noqa: PLC0415
    stocks = get_active_top_stocks()
    if not stocks:
        return []
    return _fetch_all_stocks_trades(stocks)


# ── Cell / sheet helpers ───────────────────────────────────────────────────────

def _set_cell(ws, row, col, value, bold=False, color=None, fill=None,
              align="center", border=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=color or _DARK, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if fill:
        cell.fill = fill
    if border:
        cell.border = _BORDER
    return cell


def _style_header_row(ws, row, ncols, fill=None):
    fill = fill or _HEADER_FILL
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = Font(bold=True, color=_WHITE, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER


def _autofit(ws):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 40)


def _pnl_fill(pnl):
    return _WIN_FILL if pnl >= 0 else _LOSS_FILL


def _pnl_color(pnl):
    return _GREEN if pnl >= 0 else _RED


def _kpi_row(ws, row, kpi_list):
    """Write a label row then a value row for a list of (label, value) pairs."""
    ws.row_dimensions[row].height = 16
    ws.row_dimensions[row + 1].height = 22
    for c, (lbl, val) in enumerate(kpi_list, start=1):
        _set_cell(ws, row, c, lbl, bold=True, fill=_SUB_FILL, color=_WHITE)
        is_pnl = "P&L" in lbl or "pnl" in lbl.lower()
        is_cap = "Capital" in lbl
        if is_cap:
            f   = _PURPLE_FILL
            col = "FF6C3483"
        elif is_pnl and isinstance(val, (int, float)):
            f   = _pnl_fill(val)
            col = _pnl_color(val)
        else:
            f   = _ALT_FILL
            col = _DARK
        _set_cell(ws, row + 1, c, val, bold=(is_pnl or is_cap), fill=f, color=col)


# ── Strategy-breakdown table ───────────────────────────────────────────────────

def _write_strategy_table(ws, start_row, trades):
    headers = [
        "Strategy", "Trades", "Winners", "Losers",
        "Win Rate %", "Total P&L", "Avg Win", "Avg Loss", "Best Trade",
    ]
    _style_header_row(ws, start_row, len(headers))
    for c, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=c, value=h)

    strat_pnl_map = {}
    for i, strat in enumerate(_ALL_STRATEGIES):
        r = start_row + 1 + i
        ws.row_dimensions[r].height = 16
        st = [t for t in trades if t.get("strategy") == strat]
        spnl  = round(sum(t["pnl"] for t in st), 2)
        sw    = sum(1 for t in st if t["pnl"] > 0)
        sl    = sum(1 for t in st if t["pnl"] < 0)
        swr   = round(sw / len(st) * 100, 1) if st else 0
        s_aw  = round(sum(t["pnl"] for t in st if t["pnl"] > 0) / sw, 2) if sw else 0
        s_al  = round(sum(t["pnl"] for t in st if t["pnl"] < 0) / sl, 2) if sl else 0
        best  = round(max((t["pnl"] for t in st), default=0), 2)
        row_fill = _ALT_FILL if i % 2 == 0 else _WHITE_FILL
        strat_pnl_map[strat] = spnl

        for c, v in enumerate(
            [strat, len(st), sw, sl, swr, spnl, s_aw, s_al, best], start=1
        ):
            f   = _pnl_fill(spnl) if c == 6 else row_fill
            col = _pnl_color(spnl) if c == 6 else _DARK
            _set_cell(ws, r, c, v, bold=(c == 6), fill=f, color=col,
                      align="left" if c == 1 else "center")

    return start_row + 1 + len(_ALL_STRATEGIES), strat_pnl_map


def _add_strategy_bar_chart(ws, data_start_row, anchor_cell):
    chart = BarChart()
    chart.type  = "col"
    chart.title = "Strategy P&L Comparison"
    chart.y_axis.title = "P&L"
    chart.x_axis.title = "Strategy"
    chart.style  = 10
    chart.width  = 22
    chart.height = 12
    data_end_row = data_start_row + len(_ALL_STRATEGIES) - 1
    chart.add_data(
        Reference(ws, min_col=6, max_col=6, min_row=data_start_row, max_row=data_end_row),
        titles_from_data=False,
    )
    chart.series[0].title = SeriesLabel(v="P&L")
    chart.set_categories(
        Reference(ws, min_col=1, max_col=1, min_row=data_start_row, max_row=data_end_row)
    )
    ws.add_chart(chart, anchor_cell)


# ── Stock-breakdown table ──────────────────────────────────────────────────────

def _write_stock_table(ws, start_row, trades):
    """Per-stock P&L summary. Returns the row after the last data row."""
    stock_names = sorted(set(t.get("stock", "") for t in trades if t.get("stock")))
    if not stock_names:
        return start_row

    headers = ["Stock", "Trades", "Winners", "Losers", "Win Rate %", "Total P&L", "Capital Invested"]
    _style_header_row(ws, start_row, len(headers))
    for c, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=c, value=h)

    for i, stock in enumerate(stock_names):
        r = start_row + 1 + i
        ws.row_dimensions[r].height = 16
        st = [t for t in trades if t.get("stock") == stock]
        spnl = round(sum(t["pnl"] for t in st), 2)
        sw   = sum(1 for t in st if t["pnl"] > 0)
        sl   = sum(1 for t in st if t["pnl"] < 0)
        swr  = round(sw / len(st) * 100, 1) if st else 0
        cap  = round(sum(float(t.get("entry", 0) or 0) for t in st), 2)
        rf   = _ALT_FILL if i % 2 == 0 else _WHITE_FILL

        for c, v in enumerate([stock, len(st), sw, sl, swr, spnl, cap], start=1):
            f   = _pnl_fill(spnl) if c == 6 else (_PURPLE_FILL if c == 7 else rf)
            col = _pnl_color(spnl) if c == 6 else ("FF6C3483" if c == 7 else _DARK)
            _set_cell(ws, r, c, v, bold=(c in (6, 7)), fill=f, color=col,
                      align="left" if c == 1 else "center")

    return start_row + 1 + len(stock_names)


# ── Trade-detail table ─────────────────────────────────────────────────────────

def _write_trade_table(ws, start_row, trades):
    """
    Write a full trade table with cumulative P&L and an equity curve.
    Returns the row after the last data row.
    """
    headers = [
        "Date", "Time", "Stock", "Strategy", "Signal", "Type",
        "Entry", "Target", "Stop Loss", "Exit", "Status", "P&L", "Cumulative P&L",
    ]
    ws.row_dimensions[start_row].height = 16
    _style_header_row(ws, start_row, len(headers))
    for c, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=c, value=h)

    cumulative = 0.0
    equity_rows = []
    for i, trade in enumerate(sorted(trades, key=lambda t: t.get("time") or "")):
        r = start_row + 1 + i
        ws.row_dimensions[r].height = 15
        pnl = trade.get("pnl", 0)
        cumulative = round(cumulative + pnl, 2)
        equity_rows.append(r)

        entry_time = trade.get("time")
        date_str = time_str = ""
        if hasattr(entry_time, "strftime"):
            date_str = entry_time.strftime("%Y-%m-%d")
            time_str = entry_time.strftime("%H:%M")
        elif isinstance(entry_time, str):
            date_str = entry_time[:10]
            time_str = entry_time[11:16] if len(entry_time) > 10 else ""

        row_fill = _ALT_FILL if i % 2 == 0 else _WHITE_FILL
        vals = [
            date_str, time_str,
            trade.get("stock", ""), trade.get("strategy", ""),
            trade.get("signal", ""), trade.get("type", ""),
            trade.get("entry"), trade.get("target"), trade.get("stop_loss"),
            trade.get("exit_price"), trade.get("status", ""),
            round(pnl, 2), cumulative,
        ]
        fills  = [row_fill] * 11 + [_pnl_fill(pnl), _pnl_fill(cumulative)]
        colors = [_DARK] * 11 + [_pnl_color(pnl), _pnl_color(cumulative)]
        bolds  = [False] * 11 + [True, True]
        aligns = ["left"] * 6 + ["center"] * 7

        for c, (v, f, col, b, a) in enumerate(
            zip(vals, fills, colors, bolds, aligns), start=1
        ):
            _set_cell(ws, r, c, v, bold=b, fill=f, color=col, align=a)

    last_data_row = start_row + len(trades)

    # Equity curve (column 13 = Cumulative P&L)
    if equity_rows:
        chart = LineChart()
        chart.title = "Equity Curve"
        chart.y_axis.title = "Cumulative P&L"
        chart.style = 10
        chart.width = 22
        chart.height = 12
        chart.add_data(
            Reference(ws, min_col=13, max_col=13,
                      min_row=equity_rows[0], max_row=equity_rows[-1]),
            titles_from_data=False,
        )
        chart.series[0].title = SeriesLabel(v="Equity")
        ws.add_chart(chart, f"O{start_row}")

    return last_data_row + 1


# ── Daily P&L summary table (used in weekly summary) ─────────────────────────

def _write_day_summary_table(ws, start_row, trades_by_date):
    headers = ["Date", "Trades", "Winners", "Losers", "Win Rate %", "P&L"]
    _style_header_row(ws, start_row, len(headers))
    for c, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=c, value=h)

    for i, date in enumerate(sorted(trades_by_date.keys())):
        r = start_row + 1 + i
        ws.row_dimensions[r].height = 16
        dt = trades_by_date[date]
        dpnl = round(sum(t["pnl"] for t in dt), 2)
        dw   = sum(1 for t in dt if t["pnl"] > 0)
        dl   = sum(1 for t in dt if t["pnl"] < 0)
        dwr  = round(dw / len(dt) * 100, 1) if dt else 0
        rf   = _ALT_FILL if i % 2 == 0 else _WHITE_FILL
        for c, v in enumerate([date, len(dt), dw, dl, dwr, dpnl], start=1):
            f   = _pnl_fill(dpnl) if c == 6 else rf
            col = _pnl_color(dpnl) if c == 6 else _DARK
            _set_cell(ws, r, c, v, bold=(c == 6), fill=f, color=col,
                      align="left" if c == 1 else "center")

    return start_row + 1 + len(trades_by_date)


# ── Title banner helper ────────────────────────────────────────────────────────

def _title_banner(ws, merge_range, text):
    ws.merge_cells(merge_range)
    first_col = merge_range.split(":")[0]
    cell = ws[first_col]
    cell.value = text
    cell.font = Font(bold=True, size=14, color=_WHITE)
    cell.fill = _SUMMARY_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[int("".join(filter(str.isdigit, first_col)))].height = 30


def _section_header(ws, row, merge_to_col, text):
    ws.merge_cells(f"A{row}:{get_column_letter(merge_to_col)}{row}")
    cell = ws.cell(row=row, column=1, value=text)
    cell.font  = Font(bold=True, size=11, color=_WHITE)
    cell.fill  = _SUB_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18


# ══════════════════════════════════════════════════════════════════════════════
# DAILY REPORT (Mon–Thu):  Summary sheet + one sheet per strategy
# ══════════════════════════════════════════════════════════════════════════════

def _build_daily_summary_sheet(wb, today_trades, report_date):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    stock_list = sorted(set(t.get("stock", "") for t in today_trades if t.get("stock")))
    stocks_label = ", ".join(stock_list[:5]) + ("…" if len(stock_list) > 5 else "")
    _title_banner(ws, "A1:J1", f"Daily Backtest P&L — Top Stocks  |  {report_date}")

    total_pnl    = round(sum(t["pnl"] for t in today_trades), 2)
    total_trades = len(today_trades)
    winners      = sum(1 for t in today_trades if t["pnl"] > 0)
    losers       = sum(1 for t in today_trades if t["pnl"] < 0)
    win_rate     = round(winners / total_trades * 100, 1) if total_trades else 0
    avg_win      = round(sum(t["pnl"] for t in today_trades if t["pnl"] > 0) / winners, 2) if winners else 0
    avg_loss     = round(sum(t["pnl"] for t in today_trades if t["pnl"] < 0) / losers, 2) if losers else 0
    capital      = round(sum(float(t.get("entry", 0) or 0) for t in today_trades), 2)

    _kpi_row(ws, 3, [
        ("Total P&L", total_pnl), ("Trades", total_trades),
        ("Winners", winners), ("Losers", losers),
        ("Win Rate %", win_rate), ("Avg Win", avg_win),
        ("Avg Loss", avg_loss), ("Capital Invested", capital),
    ])

    # Stock breakdown section
    _section_header(ws, 6, 7, f"Stock Breakdown  ({len(stock_list)} active stocks)")
    after_stock = _write_stock_table(ws, 7, today_trades)

    # Strategy breakdown section
    _section_header(ws, after_stock + 1, 9, "Strategy Analysis")
    _, strat_pnl_map = _write_strategy_table(ws, after_stock + 2, today_trades)
    _add_strategy_bar_chart(ws, after_stock + 3, "K3")

    _autofit(ws)


def _build_daily_strategy_sheet(wb, strat_name, trades):
    """One sheet per strategy, today's trades only."""
    ws = wb.create_sheet(title=strat_name.replace("/", "-")[:31])
    ws.sheet_view.showGridLines = False

    _title_banner(ws, "A1:M1", f"{strat_name} — Today's Trades")

    pnl_total = round(sum(t["pnl"] for t in trades), 2)
    winners   = sum(1 for t in trades if t["pnl"] > 0)
    losers    = sum(1 for t in trades if t["pnl"] < 0)
    win_rate  = round(winners / len(trades) * 100, 1) if trades else 0
    capital   = round(sum(float(t.get("entry", 0) or 0) for t in trades), 2)

    _kpi_row(ws, 2, [
        ("Trades", len(trades)), ("Winners", winners),
        ("Losers", losers), ("Win Rate %", win_rate),
        ("Total P&L", pnl_total), ("Capital Invested", capital),
    ])

    if trades:
        _write_trade_table(ws, 5, trades)
    else:
        ws.cell(row=5, column=1, value="No trades today for this strategy.").font = Font(
            italic=True, color=_DARK
        )

    _autofit(ws)


def build_daily_excel(all_trades, today_str, report_date):
    """Excel for Mon–Thu: today's trades only."""
    today_trades = [t for t in all_trades if t.get("trade_date") == today_str]
    wb = openpyxl.Workbook()
    _build_daily_summary_sheet(wb, today_trades, report_date)
    for strat in _ALL_STRATEGIES:
        _build_daily_strategy_sheet(wb, strat, [t for t in today_trades if t.get("strategy") == strat])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), today_trades


# ══════════════════════════════════════════════════════════════════════════════
# WEEKLY REPORT (Friday):  Weekly Summary sheet + one sheet per trading day
# ══════════════════════════════════════════════════════════════════════════════

def _build_weekly_summary_sheet(wb, all_trades, week_label):
    ws = wb.active
    ws.title = "Weekly Summary"
    ws.sheet_view.showGridLines = False

    _title_banner(ws, "A1:J1", f"Weekly Backtest P&L — Top Stocks  |  {week_label}")

    total_pnl    = round(sum(t["pnl"] for t in all_trades), 2)
    total_trades = len(all_trades)
    winners      = sum(1 for t in all_trades if t["pnl"] > 0)
    losers       = sum(1 for t in all_trades if t["pnl"] < 0)
    win_rate     = round(winners / total_trades * 100, 1) if total_trades else 0
    avg_win      = round(sum(t["pnl"] for t in all_trades if t["pnl"] > 0) / winners, 2) if winners else 0
    avg_loss     = round(sum(t["pnl"] for t in all_trades if t["pnl"] < 0) / losers, 2) if losers else 0
    capital      = round(sum(float(t.get("entry", 0) or 0) for t in all_trades), 2)

    _kpi_row(ws, 3, [
        ("Total P&L", total_pnl), ("Trades", total_trades),
        ("Winners", winners), ("Losers", losers),
        ("Win Rate %", win_rate), ("Avg Win", avg_win),
        ("Avg Loss", avg_loss), ("Capital Invested", capital),
    ])

    # Stock breakdown
    _section_header(ws, 6, 7, "Stock Breakdown")
    after_stock = _write_stock_table(ws, 7, all_trades)

    # Day-wise breakdown
    _section_header(ws, after_stock + 1, 6, "Day-wise P&L")
    trades_by_date = defaultdict(list)
    for t in all_trades:
        trades_by_date[t.get("trade_date", "Unknown")].append(t)
    after_day_table = _write_day_summary_table(ws, after_stock + 2, trades_by_date)

    # Strategy breakdown
    strat_section_row = after_day_table + 2
    _section_header(ws, strat_section_row, 9, "Strategy Analysis (Full Week)")
    strat_data_start = strat_section_row + 1
    _write_strategy_table(ws, strat_data_start, all_trades)
    _add_strategy_bar_chart(ws, strat_data_start + 1, "K3")

    _autofit(ws)


def _build_day_sheet(wb, date_str, day_trades):
    """One sheet per trading day with all strategies' trades and analysis."""
    ws = wb.create_sheet(title=date_str)
    ws.sheet_view.showGridLines = False

    _title_banner(ws, "A1:M1", f"Trades — {date_str}")

    pnl_total = round(sum(t["pnl"] for t in day_trades), 2)
    winners   = sum(1 for t in day_trades if t["pnl"] > 0)
    losers    = sum(1 for t in day_trades if t["pnl"] < 0)
    win_rate  = round(winners / len(day_trades) * 100, 1) if day_trades else 0
    capital   = round(sum(float(t.get("entry", 0) or 0) for t in day_trades), 2)

    _kpi_row(ws, 2, [
        ("Trades", len(day_trades)), ("Winners", winners),
        ("Losers", losers), ("Win Rate %", win_rate),
        ("Total P&L", pnl_total), ("Capital Invested", capital),
    ])

    # Stock breakdown
    _section_header(ws, 5, 7, "Stock Breakdown")
    after_stock = _write_stock_table(ws, 6, day_trades)

    # Per-strategy mini-summary
    _section_header(ws, after_stock + 1, 6, "Strategy Breakdown")
    strat_headers = ["Strategy", "Trades", "Winners", "Losers", "Win Rate %", "P&L"]
    _style_header_row(ws, after_stock + 2, len(strat_headers))
    for c, h in enumerate(strat_headers, start=1):
        ws.cell(row=after_stock + 2, column=c, value=h)

    for i, strat in enumerate(_ALL_STRATEGIES):
        r = after_stock + 3 + i
        ws.row_dimensions[r].height = 15
        st = [t for t in day_trades if t.get("strategy") == strat]
        spnl = round(sum(t["pnl"] for t in st), 2)
        sw   = sum(1 for t in st if t["pnl"] > 0)
        sl   = sum(1 for t in st if t["pnl"] < 0)
        swr  = round(sw / len(st) * 100, 1) if st else 0
        rf   = _ALT_FILL if i % 2 == 0 else _WHITE_FILL
        for c, v in enumerate([strat, len(st), sw, sl, swr, spnl], start=1):
            f   = _pnl_fill(spnl) if c == 6 else rf
            col = _pnl_color(spnl) if c == 6 else _DARK
            _set_cell(ws, r, c, v, bold=(c == 6), fill=f, color=col,
                      align="left" if c == 1 else "center")

    # Full trade table
    trade_header_row = after_stock + 3 + len(_ALL_STRATEGIES) + 2
    _section_header(ws, trade_header_row, 13, "All Trades")

    if day_trades:
        _write_trade_table(ws, trade_header_row + 1, day_trades)
    else:
        ws.cell(row=trade_header_row + 1, column=1,
                value="No completed trades this day.").font = Font(italic=True, color=_DARK)

    _autofit(ws)


def build_weekly_excel(all_trades, week_label):
    """Excel for Friday: full week, one sheet per trading day."""
    trades_by_date = defaultdict(list)
    for t in all_trades:
        trades_by_date[t.get("trade_date", "Unknown")].append(t)

    wb = openpyxl.Workbook()
    _build_weekly_summary_sheet(wb, all_trades, week_label)
    for date in sorted(trades_by_date.keys()):
        _build_day_sheet(wb, date, trades_by_date[date])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Email sender ───────────────────────────────────────────────────────────────

def _send_email(subject, body, attachment_bytes, filename):
    mail_enabled = os.getenv("MAIL_ENABLED", "True").strip().lower() not in ("false", "0", "no")
    if not mail_enabled:
        print(f"  [email_report] disabled — skipping: {subject}")
        return

    gmail_user   = os.getenv("GMAIL_USERNAME", "")
    app_password = os.getenv("GMAIL_APP_PASSWORD", "")
    receiver_raw = os.getenv("RECEIVER_EMAIL", "")
    receivers    = [e.strip() for e in receiver_raw.split(",") if e.strip()]

    if not gmail_user or not app_password or not receivers:
        print("  [email_report] Missing email credentials or receiver. Skipping.")
        return

    msg = MIMEMultipart()
    msg["From"]    = gmail_user
    msg["To"]      = ", ".join(receivers)
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "html"))

    part = MIMEBase("application", "octet-stream")
    part.set_payload(attachment_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(gmail_user, app_password)
        server.sendmail(gmail_user, receivers, msg.as_string())

    print(f"  [email_report] Report sent to {receivers}")


# ── Public entry point ─────────────────────────────────────────────────────────

def send_backtest_email_report():
    """
    Called every minute by the scheduler.
    Fires once at 4:00–4:10 PM IST on trading days.
    Mon–Thu → daily report (today only).
    Friday  → weekly report (full 5-day window).
    """
    now = now_ist()
    today_str = now.strftime("%Y-%m-%d")
    report_key = f"backtest_email_{today_str}"

    if not (now.hour == 16 and 0 <= now.minute <= 10):
        return
    if not _is_trading_day(now):
        return
    if _is_already_sent(report_key):
        return

    is_friday = now.weekday() == 4
    print(f"  [email_report] Building {'weekly' if is_friday else 'daily'} report for {today_str}…")

    try:
        all_trades = _fetch_all_stocks_for_report()

        # Collect stock names for the email body
        stock_names = sorted(set(t.get("stock", "") for t in all_trades if t.get("stock")))
        stocks_html = ", ".join(f"<strong>{s}</strong>" for s in stock_names[:10])
        if len(stock_names) > 10:
            stocks_html += f" + {len(stock_names) - 10} more"

        capital = round(sum(float(t.get("entry", 0) or 0) for t in all_trades), 2)

        if is_friday:
            week_label  = f"Week ending {now.strftime('%d %b %Y')}"
            excel_bytes = build_weekly_excel(all_trades, week_label)
            filename    = f"backtest_weekly_{today_str}.xlsx"
            total_pnl   = round(sum(t.get("pnl", 0) for t in all_trades), 2)
            sign        = "+" if total_pnl >= 0 else ""
            report_type = "Weekly"
            extra_note  = "Each sheet covers one trading day (all strategies &amp; stocks)."
        else:
            report_date  = now.strftime("%A, %d %b %Y")
            today_trades_list = [t for t in all_trades if t.get("trade_date") == today_str]
            excel_bytes, today_trades_list = build_daily_excel(all_trades, today_str, report_date)
            filename    = f"backtest_daily_{today_str}.xlsx"
            total_pnl   = round(sum(t.get("pnl", 0) for t in today_trades_list), 2)
            sign        = "+" if total_pnl >= 0 else ""
            report_type = "Daily"
            extra_note  = "Each sheet covers one strategy (today's trades across all stocks)."
            capital     = round(sum(float(t.get("entry", 0) or 0) for t in today_trades_list), 2)

        report_date_label = (
            f"Week ending {now.strftime('%d %b %Y')}" if is_friday
            else now.strftime("%A, %d %b %Y")
        )
        pnl_color = "#1e8449" if total_pnl >= 0 else "#c0392b"

        body = f"""
<html><body style="font-family:Arial,sans-serif;color:#2c3e50;">
<h2 style="color:#1a5276;">{report_type} Backtest P&amp;L Report — Top Stocks</h2>
<p><strong>Period:</strong> {report_date_label}</p>
<p><strong>Stocks covered:</strong> {stocks_html}</p>
<table style="border-collapse:collapse;margin:12px 0;">
  <tr>
    <td style="padding:6px 16px 6px 0;color:#555;">Overall P&amp;L</td>
    <td style="color:{pnl_color};font-size:20px;font-weight:bold;">{sign}{total_pnl:.2f}</td>
  </tr>
  <tr>
    <td style="padding:6px 16px 6px 0;color:#555;">Capital Invested</td>
    <td style="color:#6c3483;font-size:18px;font-weight:bold;">₹{capital:,.2f}</td>
  </tr>
</table>
<p>Please find the detailed Excel report attached. It contains:</p>
<ul>
  <li><strong>{'Weekly Summary' if is_friday else 'Summary'}</strong> sheet — overall KPIs,
      stock breakdown, {'day-wise breakdown, ' if is_friday else ''}strategy comparison &amp; bar chart</li>
  <li>{extra_note}</li>
</ul>
<hr/>
<p style="font-size:11px;color:#7f8c8d;">
  Auto-generated by AlgoTrading platform at 4:00 PM IST. Do not reply to this email.
</p>
</body></html>
"""
        _send_email(
            subject=f"[AlgoTrading] {report_type} P&L Report — {today_str}  ({sign}{total_pnl:.2f})",
            body=body,
            attachment_bytes=excel_bytes,
            filename=filename,
        )
        _mark_sent(report_key)

    except Exception:
        print(f"  [email_report] Failed:\n{traceback.format_exc()}")
